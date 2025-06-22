#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ステップ4: Excel出力機能
ファイルパスと抽出した画像をExcelファイルに出力する
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
import sys
import io
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def resize_image_for_excel(image: Image.Image, target_size: int = 100) -> Image.Image:
    """
    Excelセル用に画像をリサイズ（アスペクト比維持）
    
    Args:
        image (Image.Image): 元の画像
        target_size (int): 目標サイズ（px）
        
    Returns:
        Image.Image: リサイズされた画像
    """
    # アスペクト比を維持してtarget_size以内に収まるようにリサイズ
    image.thumbnail((target_size, target_size), Image.Resampling.LANCZOS)
    
    # 新しい正方形の白背景画像を作成
    resized = Image.new('RGB', (target_size, target_size), (255, 255, 255))
    
    # 元画像を中央に配置
    offset = ((target_size - image.size[0]) // 2, (target_size - image.size[1]) // 2)
    
    # 透明度がある場合の処理
    if image.mode == 'RGBA':
        resized.paste(image, offset, image)
    else:
        resized.paste(image, offset)
    
    return resized

def save_image_to_bytes(image: Image.Image) -> io.BytesIO:
    """
    PIL ImageをBytesIOに変換してExcelで使用可能にする
    
    Args:
        image (Image.Image): PIL Image
        
    Returns:
        io.BytesIO: バイト形式の画像データ
    """
    img_buffer = io.BytesIO()
    image.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    return img_buffer

def create_excel_with_images(file_data: List[Dict[str, Any]], output_path: str = "result.xlsx") -> None:
    """
    ファイル情報と画像データをExcelファイルに出力
    
    Args:
        file_data (List[Dict[str, Any]]): ファイルデータのリスト
            各辞書には以下のキーが含まれる:
            - 'file_path': ファイルの絶対パス
            - 'images': 画像情報のリスト
        output_path (str): 出力ファイルパス
    """
    try:
        # 新しいワークブックとワークシートを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "ファイルと画像の一覧"
        
        # ヘッダーを設定
        headers = ["ファイルパス", "画像1", "画像2", "画像3"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 列幅の設定
        ws.column_dimensions['A'].width = 60  # ファイルパス列
        for col in range(2, 5):  # 画像列（B, C, D）
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15  # 100px ≈ 15文字分
        
        # 行の高さを設定（100px ≈ 75ポイント）
        current_row = 2
        
        for file_info in file_data:
            # ファイルパスを設定
            file_path_cell = ws.cell(row=current_row, column=1, value=str(file_info['file_path']))
            file_path_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 最大3枚の画像まで表示
            images = file_info.get('images', [])
            max_images = min(len(images), 3)
            
            for img_index in range(max_images):
                img_info = images[img_index]
                
                try:
                    # 画像をExcel用にリサイズ
                    resized_image = resize_image_for_excel(img_info['image'], 100)
                    
                    # 画像をBytesIOに変換
                    img_buffer = save_image_to_bytes(resized_image)
                    
                    # Excelに画像を挿入
                    excel_img = ExcelImage(img_buffer)
                    
                    # 画像のサイズを調整（ピクセル単位）
                    excel_img.width = 100
                    excel_img.height = 100
                    
                    # 画像をセルに配置（B列=2, C列=3, D列=4）
                    target_column = img_index + 2
                    cell_position = f"{get_column_letter(target_column)}{current_row}"
                    ws.add_image(excel_img, cell_position)
                    
                except Exception as e:
                    # 画像の処理に失敗した場合、エラーメッセージをセルに記入
                    error_cell = ws.cell(row=current_row, column=img_index + 2, value=f"画像エラー: {str(e)[:20]}...")
                    error_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 行の高さを100pxに設定（約75ポイント）
            ws.row_dimensions[current_row].height = 75
            
            current_row += 1
        
        # Excelファイルを保存
        wb.save(output_path)
        print(f"Excelファイルが正常に作成されました: {output_path}")
        
    except Exception as e:
        raise Exception(f"Excel出力中にエラーが発生しました: {e}")

def test_excel_output() -> None:
    """
    Excel出力機能のテスト
    """
    print("=== Excel出力機能テスト ===")
    
    # step1, step2, step3の機能をインポート
    from step1_file_crawler import find_files
    from step2_docx_image_extractor import extract_images_from_docx
    from step3_pdf_image_extractor import extract_images_from_pdf
    
    try:
        # ファイルを検索
        files = find_files("target")
        
        if not files:
            print("テスト対象のファイルが見つかりません。")
            return
        
        file_data = []
        
        # 最初の数ファイルのみでテスト（時間短縮のため）
        test_files = files[:3]  # 最初の3ファイル
        
        print(f"テスト対象ファイル数: {len(test_files)}")
        
        for file_path in test_files:
            print(f"処理中: {file_path.name}")
            
            images = []
            
            # ファイル形式に応じて画像を抽出
            if file_path.suffix.lower() == '.docx':
                try:
                    images = extract_images_from_docx(file_path)
                except Exception as e:
                    print(f"警告: {file_path.name} の処理中にエラー: {e}")
                    
            elif file_path.suffix.lower() == '.pdf':
                try:
                    images = extract_images_from_pdf(file_path)
                except Exception as e:
                    print(f"警告: {file_path.name} の処理中にエラー: {e}")
            
            # ファイル情報を追加
            file_info = {
                'file_path': file_path.absolute(),
                'images': images
            }
            
            file_data.append(file_info)
            print(f"  → 画像 {len(images)} 枚を抽出")
        
        # Excelファイルを作成
        print("\nExcelファイルを作成中...")
        create_excel_with_images(file_data, "test_result.xlsx")
        
        # 統計情報を表示
        total_images = sum(len(f['images']) for f in file_data)
        print(f"\n=== 統計情報 ===")
        print(f"処理ファイル数: {len(file_data)}")
        print(f"総抽出画像数: {total_images}")
        
        print(f"\n=== テスト完了 ===")
        print("test_result.xlsx を確認してください。")
        
    except Exception as e:
        print(f"テスト実行エラー: {e}")
        sys.exit(1)

def main():
    """メイン関数 - テスト実行"""
    test_excel_output()

if __name__ == "__main__":
    main() 
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
改良版メインプログラム: 文書画像抽出システム
targetディレクトリを再帰的にクロールし、.docxと.pdfファイルから全ての画像を抽出してresult.xlsxに出力
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
import sys
import io
import time
from PIL import Image
from docx import Document
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ===== ファイルクロール機能 =====
def find_files(target_dir: str, extensions: tuple = ('.docx', '.pdf')) -> List[Path]:
    """指定ディレクトリを再帰的に検索し、指定拡張子のファイルを収集"""
    target_path = Path(target_dir)
    
    if not target_path.exists():
        raise FileNotFoundError(f"指定されたディレクトリが見つかりません: {target_dir}")
    
    if not target_path.is_dir():
        raise NotADirectoryError(f"指定されたパスはディレクトリではありません: {target_dir}")
    
    found_files = []
    
    for extension in extensions:
        pattern = f"**/*{extension}"
        files = list(target_path.rglob(pattern))
        found_files.extend(files)
    
    # 重複を除去し、パスでソート
    found_files = sorted(list(set(found_files)))
    
    return found_files

# ===== 画像抽出機能（.docx） =====
def extract_images_from_docx(docx_path: Path) -> List[Dict[str, Any]]:
    """.docxファイルから画像を抽出"""
    if not docx_path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {docx_path}")
    
    try:
        doc = Document(docx_path)
        images = []
        rels = doc.part.rels
        image_index = 0
        
        for rel_id, rel in rels.items():
            if "image" in rel.target_part.content_type:
                try:
                    image_data = rel.target_part.blob
                    image_stream = io.BytesIO(image_data)
                    pil_image = Image.open(image_stream)
                    
                    image_info = {
                        'image': pil_image.copy(),
                        'format': pil_image.format or 'UNKNOWN',
                        'size': pil_image.size,
                        'index': image_index,
                        'source_type': 'docx'
                    }
                    
                    images.append(image_info)
                    image_index += 1
                    image_stream.close()
                    
                except Exception as e:
                    print(f"警告: DOCX画像 {rel_id} の処理中にエラー: {e}")
                    continue
                    
    except Exception as e:
        raise Exception(f"DOCX ファイルの処理中にエラーが発生しました: {e}")
    
    return images

# ===== 画像抽出機能（.pdf） =====
def extract_images_from_pdf(pdf_path: Path) -> List[Dict[str, Any]]:
    """.pdfファイルから画像を抽出"""
    if not pdf_path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {pdf_path}")
    
    try:
        pdf_doc = fitz.open(pdf_path)
        images = []
        image_index = 0
        
        for page_num in range(len(pdf_doc)):
            page = pdf_doc[page_num]
            image_list = page.get_images(full=True)
            
            for img_index, img in enumerate(image_list):
                try:
                    xref = img[0]
                    base_image = pdf_doc.extract_image(xref)
                    image_bytes = base_image["image"]
                    image_ext = base_image["ext"]
                    
                    image_stream = io.BytesIO(image_bytes)
                    pil_image = Image.open(image_stream)
                    
                    # カラーモード正規化
                    if pil_image.mode not in ('RGB', 'RGBA'):
                        if pil_image.mode == 'CMYK':
                            pil_image = pil_image.convert('RGB')
                        elif pil_image.mode in ('P', 'L'):
                            pil_image = pil_image.convert('RGB')
                    
                    image_info = {
                        'image': pil_image.copy(),
                        'format': image_ext.upper(),
                        'size': pil_image.size,
                        'index': image_index,
                        'page': page_num + 1,
                        'source_type': 'pdf'
                    }
                    
                    images.append(image_info)
                    image_index += 1
                    image_stream.close()
                    
                except Exception as e:
                    print(f"警告: PDF画像 (ページ {page_num + 1}) の処理中にエラー: {e}")
                    continue
        
        pdf_doc.close()
        
    except Exception as e:
        raise Exception(f"PDF ファイルの処理中にエラーが発生しました: {e}")
    
    return images

# ===== 画像リサイズ機能 =====
def resize_image_for_excel(image: Image.Image, target_size: int = 100) -> Image.Image:
    """画像を100x100pxにリサイズ（アスペクト比維持）"""
    image.thumbnail((target_size, target_size), Image.Resampling.LANCZOS)
    
    resized = Image.new('RGB', (target_size, target_size), (255, 255, 255))
    offset = ((target_size - image.size[0]) // 2, (target_size - image.size[1]) // 2)
    
    if image.mode == 'RGBA':
        resized.paste(image, offset, image)
    else:
        resized.paste(image, offset)
    
    return resized

def save_image_to_bytes(image: Image.Image) -> io.BytesIO:
    """PIL ImageをBytesIOに変換"""
    img_buffer = io.BytesIO()
    image.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    return img_buffer

# ===== Excel出力機能（改良版） =====
def create_excel_with_images(file_data: List[Dict[str, Any]], output_path: str = "result.xlsx") -> None:
    """ファイル情報と画像データをExcelファイルに出力（全画像対応）"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "ファイルと画像の一覧"
        
        # 最大画像数を取得
        max_images = max((len(f.get('images', [])) for f in file_data), default=0)
        
        # ヘッダー設定（シンプル版）
        ws.cell(row=1, column=1, value="ファイルパス").font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        
        # 列幅設定
        ws.column_dimensions['A'].width = 60  # ファイルパス列
        
        # 画像列の幅設定（B列以降）
        for col in range(2, 2 + max_images):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
        
        current_row = 2
        
        for file_info in file_data:
            # ファイルパス設定
            file_path_cell = ws.cell(row=current_row, column=1, value=str(file_info['file_path']))
            file_path_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 全ての画像を設定
            images = file_info.get('images', [])
            
            for img_index, img_info in enumerate(images):
                try:
                    resized_image = resize_image_for_excel(img_info['image'], 100)
                    img_buffer = save_image_to_bytes(resized_image)
                    excel_img = ExcelImage(img_buffer)
                    excel_img.width = 100
                    excel_img.height = 100
                    
                    # B列から順番に配置（B=2, C=3, D=4, E=5...）
                    target_column = img_index + 2
                    cell_position = f"{get_column_letter(target_column)}{current_row}"
                    ws.add_image(excel_img, cell_position)
                    
                except Exception as e:
                    # 画像エラーの場合もセルに記録
                    error_cell = ws.cell(row=current_row, column=img_index + 2, 
                                       value=f"エラー")
                    error_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 行の高さ設定（100px ≈ 75ポイント）
            ws.row_dimensions[current_row].height = 75
            current_row += 1
        
        wb.save(output_path)
        print(f"Excelファイルが正常に作成されました: {output_path}")
        print(f"最大画像数: {max_images} 枚/ファイル")
        
    except Exception as e:
        raise Exception(f"Excel出力中にエラーが発生しました: {e}")

# ===== メイン処理 =====
def main():
    """メイン処理"""
    print("=" * 60)
    print("       改良版文書画像抽出システム")
    print("  .docx/.pdfファイルから全画像を抽出してExcel出力")
    print("=" * 60)
    
    target_directory = "target"
    output_file = "result.xlsx"
    
    start_time = time.time()
    
    try:
        # Step 1: ファイル検索
        print(f"\n🔍 ステップ1: ファイル検索中...")
        print(f"   対象ディレクトリ: {target_directory}")
        
        files = find_files(target_directory)
        
        if not files:
            print("❌ 対象のファイルが見つかりませんでした。")
            print(f"   {target_directory} ディレクトリに .docx または .pdf ファイルが存在することを確認してください。")
            return
        
        docx_files = [f for f in files if f.suffix.lower() == '.docx']
        pdf_files = [f for f in files if f.suffix.lower() == '.pdf']
        
        print(f"✅ 見つかったファイル:")
        print(f"   📄 .docx ファイル: {len(docx_files)} 個")
        print(f"   📄 .pdf ファイル: {len(pdf_files)} 個")
        print(f"   📄 総ファイル数: {len(files)} 個")
        
        # Step 2: 画像抽出処理
        print(f"\n🖼️  ステップ2: 画像抽出中...")
        
        file_data = []
        total_extracted_images = 0
        
        for i, file_path in enumerate(files, 1):
            print(f"   処理中 ({i}/{len(files)}): {file_path.name}")
            
            images = []
            
            try:
                if file_path.suffix.lower() == '.docx':
                    images = extract_images_from_docx(file_path)
                elif file_path.suffix.lower() == '.pdf':
                    images = extract_images_from_pdf(file_path)
                
                file_info = {
                    'file_path': file_path.absolute(),
                    'images': images
                }
                
                file_data.append(file_info)
                total_extracted_images += len(images)
                
                print(f"     → 画像 {len(images)} 枚を抽出")
                
            except Exception as e:
                print(f"     ❌ エラー: {e}")
                # エラーが発生しても処理を継続
                file_info = {
                    'file_path': file_path.absolute(),
                    'images': []
                }
                file_data.append(file_info)
        
        print(f"✅ 画像抽出完了: 総 {total_extracted_images} 枚")
        
        # Step 3: Excel出力
        print(f"\n📊 ステップ3: Excel出力中...")
        print(f"   出力ファイル: {output_file}")
        
        create_excel_with_images(file_data, output_file)
        
        # 完了報告
        end_time = time.time()
        elapsed_time = end_time - start_time
        
        print(f"\n🎉 処理完了!")
        print(f"   処理時間: {elapsed_time:.2f} 秒")
        print(f"   処理ファイル数: {len(files)}")
        print(f"   抽出画像数: {total_extracted_images}")
        print(f"   出力ファイル: {output_file}")
        print("\n📋 結果の確認:")
        print(f"   {output_file} を開いて結果を確認してください。")
        print(f"   A列にファイルパス、B列以降に全ての画像（100x100px）が表示されます。")
        
    except Exception as e:
        print(f"\n❌ エラーが発生しました: {e}")
        print("\n🔧 トラブルシューティング:")
        print("   1. targetディレクトリが存在することを確認してください")
        print("   2. 処理対象ファイルが破損していないか確認してください")
        print("   3. 十分な空き容量があることを確認してください")
        sys.exit(1)

if __name__ == "__main__":
    main() 